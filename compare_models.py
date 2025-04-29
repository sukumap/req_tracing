#!/usr/bin/env python3
"""
Compare the results of different models for requirement tracing.
This script compares the output Excel files from the requirement tracing tool
to evaluate the performance of different models.
"""

import pandas as pd
import sys
import os
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from collections import defaultdict

def load_excel_data(file_path):
    """Load data from an Excel file."""
    if not os.path.exists(file_path):
        print(f"Error: File {file_path} not found.")
        return None
    
    # Load the workbook to get sheet names
    workbook = load_workbook(file_path, read_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    
    # Load the main sheet (first sheet)
    main_df = pd.read_excel(file_path, sheet_name=0)
    
    # Check if there are function sheets
    function_sheets = {}
    for sheet_name in sheet_names:
        if sheet_name.startswith("Functions for REQ-"):
            function_sheets[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
    
    return {
        "main": main_df,
        "function_sheets": function_sheets
    }

def analyze_results(data):
    """Analyze the results from a single model."""
    main_df = data["main"]
    
    # Count requirements with matches
    reqs_with_matches = main_df[main_df["Matching Functions"] > 0].shape[0]
    total_reqs = main_df.shape[0]
    
    # Count total matches
    total_matches = main_df["Matching Functions"].sum()
    
    # Calculate average confidence score
    confidence_scores = []
    for _, row in main_df.iterrows():
        if row["Matching Functions"] > 0:
            # Extract confidence scores from the "Top Functions" column
            functions_text = row["Top Functions"]
            if isinstance(functions_text, str):
                # Extract confidence scores using a simple parsing approach
                parts = functions_text.split("\n")
                for part in parts:
                    if "Confidence:" in part:
                        try:
                            confidence = float(part.split("Confidence:")[1].strip().rstrip("%")) / 100
                            confidence_scores.append(confidence)
                        except:
                            pass
    
    avg_confidence = np.mean(confidence_scores) if confidence_scores else 0
    
    # Count matches by match type
    match_types = defaultdict(int)
    for _, row in main_df.iterrows():
        if row["Matching Functions"] > 0:
            functions_text = row["Top Functions"]
            if isinstance(functions_text, str):
                parts = functions_text.split("\n")
                for part in parts:
                    if "Match Type:" in part:
                        match_type = part.split("Match Type:")[1].strip()
                        match_types[match_type] += 1
    
    return {
        "reqs_with_matches": reqs_with_matches,
        "total_reqs": total_reqs,
        "match_percentage": reqs_with_matches / total_reqs * 100 if total_reqs > 0 else 0,
        "total_matches": total_matches,
        "avg_confidence": avg_confidence,
        "match_types": dict(match_types)
    }

def compare_models(model_results):
    """Compare the results from different models."""
    print("\n=== Model Comparison ===")
    
    # Print basic statistics
    print("\nBasic Statistics:")
    print(f"{'Model':<20} {'Reqs with Matches':<20} {'Match %':<10} {'Total Matches':<15} {'Avg Confidence':<15}")
    print("-" * 80)
    
    for model_name, results in model_results.items():
        print(f"{model_name:<20} {results['reqs_with_matches']}/{results['total_reqs']:<20} "
              f"{results['match_percentage']:.2f}%{'':<4} {results['total_matches']:<15} "
              f"{results['avg_confidence']:.4f}")
    
    # Print match types
    print("\nMatch Types:")
    all_match_types = set()
    for results in model_results.values():
        all_match_types.update(results["match_types"].keys())
    
    print(f"{'Model':<20}", end="")
    for match_type in sorted(all_match_types):
        print(f"{match_type:<15}", end="")
    print()
    print("-" * (20 + 15 * len(all_match_types)))
    
    for model_name, results in model_results.items():
        print(f"{model_name:<20}", end="")
        for match_type in sorted(all_match_types):
            count = results["match_types"].get(match_type, 0)
            print(f"{count:<15}", end="")
        print()

def main():
    if len(sys.argv) < 3:
        print("Usage: python compare_models.py <model1_name> <model1_excel_file> [<model2_name> <model2_excel_file> ...]")
        sys.exit(1)
    
    # Parse command line arguments
    models = {}
    for i in range(1, len(sys.argv), 2):
        if i + 1 < len(sys.argv):
            models[sys.argv[i]] = sys.argv[i + 1]
    
    # Load and analyze data for each model
    model_results = {}
    for model_name, excel_file in models.items():
        print(f"Analyzing {model_name}...")
        data = load_excel_data(excel_file)
        if data:
            model_results[model_name] = analyze_results(data)
    
    # Compare models
    if len(model_results) > 1:
        compare_models(model_results)
    else:
        print("Need at least two models to compare.")

if __name__ == "__main__":
    main()
