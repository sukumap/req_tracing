"""
Create a sample requirements Excel file for testing the requirement tracer.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

# Path to the output file
output_file = os.path.join('tests', 'data', 'sample_requirements.xlsx')

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Requirements"

# Define header style
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Set up headers
headers = ["Requirement ID", "Requirement Description", "Module Id", "Priority", "Status", "Notes"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Set column widths
ws.column_dimensions['A'].width = 15  # Requirement ID
ws.column_dimensions['B'].width = 40  # Requirement Description
ws.column_dimensions['C'].width = 15  # Module Id
ws.column_dimensions['D'].width = 10  # Priority
ws.column_dimensions['E'].width = 15  # Status
ws.column_dimensions['F'].width = 30  # Notes

# Add requirement data that matches our sample C++ files
requirements = [
    # Math requirements
    ("REQ-MATH-001", "Addition function", "Math", "High", "Implemented", "Basic arithmetic operation"),
    ("REQ-MATH-002", "Subtraction function", "Math", "High", "Implemented", "Basic arithmetic operation"),
    ("REQ-MATH-003", "Multiplication function", "Math", "High", "Implemented", "Basic arithmetic operation"),
    
    # Utility requirements
    ("REQ-UTIL-001", "String concatenation utility", "Utility", "Medium", "Implemented", "Basic string operation"),
    ("REQ-UTIL-002", "Convert string to uppercase", "Utility", "Medium", "Implemented", "String transformation utility"),
    
    # App requirements
    ("REQ-APP-001", "Application initialization", "Core", "Critical", "Implemented", "Must handle configuration loading"),
    ("REQ-APP-002", "Application shutdown", "Core", "Critical", "Implemented", "Must ensure proper resource cleanup"),
    ("REQ-APP-003", "Process command", "Core", "High", "Implemented", "Command pattern implementation"),
    
    # Main requirements
    ("REQ-MAIN-001", "Main application entry point", "Core", "Critical", "Implemented", "Coordinates all functionality"),
    
    # Additional requirements not yet implemented in code
    ("REQ-MATH-004", "Division function", "Math", "Medium", "Pending", "Basic arithmetic operation, handle divide by zero"),
    ("REQ-APP-004", "Load user preferences", "Core", "Medium", "In Progress", "Support user customization"),
    ("REQ-UTIL-003", "String splitting function", "Utility", "Low", "Planned", "Parse comma-separated values")
]

# Add data to the worksheet
for row_num, req in enumerate(requirements, 2):
    for col_num, value in enumerate(req, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True)

# Save the workbook
wb.save(output_file)
print(f"Created {output_file} successfully!")
